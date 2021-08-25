import openpyxl

inv_file = openpyxl.load_workbook("")
product_list = inv_file["Sheet1"]

print(product_list.max_row)

for product_now in range(2, product_list.max_row):
    supplier_name = product_list.cell(product_now, 4)
